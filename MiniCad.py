import tkinter as tk
from tkinter import filedialog, messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import ezdxf
import csv
import math

try:
    import openpyxl
    XLSX_OK = True
except ImportError:
    XLSX_OK = False


# -----------------------------
# Geometrijski objekti
# -----------------------------

class Line:
    def __init__(self, x1, y1, x2, y2, linetype='CONTINUOUS'):
        self.x1, self.y1 = x1, y1
        self.x2, self.y2 = x2, y2
        self.linetype = linetype

class Circle:
    def __init__(self, cx, cy, r, linetype='CONTINUOUS'):
        self.cx, self.cy, self.r = cx, cy, r
        self.linetype = linetype

class Rectangle:
    def __init__(self, x1, y1, x2, y2, linetype='CONTINUOUS'):
        self.x1, self.y1 = x1, y1
        self.x2, self.y2 = x2, y2
        self.linetype = linetype

class Point:
    def __init__(self, x, y, linetype='CONTINUOUS'):
        self.x, self.y = x, y
        self.linetype = linetype

class Arc:
    def __init__(self, cx, cy, r, angle1, angle2, linetype='CONTINUOUS'):
        self.cx, self.cy = cx, cy
        self.r = r
        self.angle1 = angle1
        self.angle2 = angle2
        self.linetype = linetype

class Text:
    def __init__(self, x, y, content, fontsize=10,
                 rotation=0.0, oblique=0.0, linetype='CONTINUOUS'):
        self.x, self.y   = x, y
        self.content     = content
        self.fontsize    = fontsize
        self.rotation    = rotation   # DXF dxf.rotation  → angle1
        self.oblique     = oblique    # DXF dxf.oblique   → angle2
        self.linetype    = linetype

class Dimension:
    """Linearna kota između dvije točke.
    offset = pomak kotne linije okomito od pravca (u jedinicama nacrta)."""
    def __init__(self, x1, y1, x2, y2, offset=5.0, linetype='CONTINUOUS'):
        self.x1, self.y1 = x1, y1
        self.x2, self.y2 = x2, y2
        self.offset   = offset
        self.linetype = linetype


# ------------------------------------------------
# Konstante i helperi za linetype
# ------------------------------------------------

LINETYPES = ['CONTINUOUS', 'DASHED', 'HIDDEN', 'DOTTED',
             'DASHDOT', 'CENTER', 'PHANTOM', 'DIVIDE', 'BORDER']

_LINESTYLE_MAP = {
    'CONTINUOUS': '-',
    'DASHED':     '--',
    'HIDDEN':     '--',
    'DOTTED':     ':',
    'DASHDOT':    '-.',
    'CENTER':     '-.',
    'PHANTOM':    (0, (5, 1, 1, 1, 1, 1)),
    'DIVIDE':     (0, (5, 1, 1, 1, 1, 1)),
    'BORDER':     (0, (5, 1, 5, 1, 1, 1)),
}

_DXF_LT_PATTERNS = {
    'DASHED':  [0.5, -0.25],
    'HIDDEN':  [0.25, -0.125],
    'DOTTED':  [0.0, -0.25],
    'DASHDOT': [0.5, -0.25, 0.0, -0.25],
    'CENTER':  [1.25, -0.25, 0.25, -0.25],
    'PHANTOM': [1.25, -0.25, 0.0, -0.25, 0.0, -0.25],
    'DIVIDE':  [0.5, -0.25, 0.0, -0.25, 0.0, -0.25],
    'BORDER':  [0.5, -0.25, 0.0, -0.25, 0.5, -0.25],
}

def _ls(lt):
    """Vrati matplotlib linestyle za DXF linetype naziv."""
    return _LINESTYLE_MAP.get((lt or 'CONTINUOUS').upper(), '-')

def _ensure_lt(doc, name):
    """Registriraj linetype u DXF dokumentu ako već nije, vrati normalizirano ime."""
    name = (name or 'CONTINUOUS').upper()
    if name not in _DXF_LT_PATTERNS:
        return 'CONTINUOUS'
    if name not in doc.linetypes:
        doc.linetypes.add(name, pattern=_DXF_LT_PATTERNS[name])
    return name

def _mtext_plain(text):
    """Pretvara MTEXT raw string (s DXF control kodovima) u plain tekst."""
    import re
    text = text.replace('\\P', '\n').replace('\\p', '\n')
    text = text.replace('\\~', ' ')
    text = text.replace('\\\\', '\x00')
    text = re.sub(r'\\[A-Za-z][^;\\]*;', '', text)
    text = re.sub(r'\\S([^^{]*)\^([^;{]*);', r'\1/\2', text)
    text = re.sub(r'[{}]', '', text)
    text = text.replace('\x00', '\\')
    return text.strip()


# -----------------------------
# Glavna aplikacija
# -----------------------------

class CADApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mini CAD")

        self.lines      = []
        self.circles    = []
        self.rectangles = []
        self.points     = []
        self.arcs       = []
        self.texts      = []
        self.dimensions = []

        self.history = []
        self.axis_range = 100

        # Pan (desni klik + drag) i scroll zoom
        self._pan_active   = False
        self._pan_press_xy = None
        self._pan_xlim     = None
        self._pan_ylim     = None
        self._pan_inv      = None
        self._custom_xlim  = None
        self._custom_ylim  = None

        # Snap
        self.snap_enabled = tk.BooleanVar(value=False)
        self.snap_grid    = 5.0

        # Background image / kalibracija
        self.bg_image        = None   # PIL Image
        self.bg_image_obj    = None   # imshow handle
        self.bg_alpha        = tk.DoubleVar(value=0.4)
        self.bg_visible      = tk.BooleanVar(value=True)
        self.cal_points      = []     # [(x, y), ...]  — max 2, u ax koordinatama
        self.cal_markers     = []     # matplotlib artists za markere
        self.cal_scale       = None   # float: cad_jedinice / ax_jedinica_slike
        self.cal_offset      = (0.0, 0.0)  # (ox, oy) pomak slike u CAD koordinatama
        self.cal_mode        = False  # True dok čekamo klikove za kalibraciju
        self._cal_click_cid  = None   # connection id za klik event

        self.create_menu()
        self.create_canvas()
        self.create_statusbar()

    # ----------------------------------------
    # GUI
    # ----------------------------------------

    def create_menu(self):
        menubar = tk.Menu(self.root)

        # FILE
        fm = tk.Menu(menubar, tearoff=0)
        fm.add_command(label="New",      command=self.new_drawing)
        fm.add_separator()
        fm.add_command(label="Open CSV", command=self.load_csv)
        fm.add_command(label="Open DXF", command=self.load_dxf)
        if XLSX_OK:
            fm.add_command(label="Open XLSX", command=self.load_xlsx)
        fm.add_separator()
        fm.add_command(label="Save CSV",  command=self.save_csv)
        fm.add_command(label="Save DXF",  command=self.save_dxf)
        fm.add_command(label="Save SVG",  command=self.save_svg)
        if XLSX_OK:
            fm.add_command(label="Save XLSX", command=self.save_xlsx)
        fm.add_command(label="Save PNG",  command=self.save_png)
        fm.add_separator()
        fm.add_command(label="Copy SVG to Clipboard", command=self.copy_svg_clipboard)
        fm.add_command(label="Copy PNG to Clipboard  Ctrl+C", command=self.copy_png_clipboard)
        fm.add_separator()
        fm.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=fm)

        # EDIT
        em = tk.Menu(menubar, tearoff=0)
        em.add_command(label="Add Line",      command=self.add_line_dialog)
        em.add_command(label="Add Circle",    command=self.add_circle_dialog)
        em.add_command(label="Add Rectangle", command=self.add_rectangle_dialog)
        em.add_command(label="Add Point",     command=self.add_point_dialog)
        em.add_command(label="Add Arc",       command=self.add_arc_dialog)
        em.add_command(label="Add Text",      command=self.add_text_dialog)
        em.add_command(label="Add Dimension", command=self.add_dimension_dialog)
        em.add_separator()
        em.add_command(label="Undo  Ctrl+Z",      command=self.undo)
        em.add_command(label="Obriši element...", command=self.delete_dialog)
        menubar.add_cascade(label="Edit", menu=em)

        # VIEW
        vm = tk.Menu(menubar, tearoff=0)
        vm.add_command(label="Refresh",        command=self.refresh_canvas)
        vm.add_command(label="Reset pogleda",  command=self.reset_view)
        vm.add_separator()
        vm.add_command(label="Raspon: 10",   command=lambda: self.set_range(10))
        vm.add_command(label="Raspon: 100",  command=lambda: self.set_range(100))
        vm.add_command(label="Raspon: 1000", command=lambda: self.set_range(1000))
        vm.add_command(label="Auto raspon",  command=self.auto_range)
        menubar.add_cascade(label="View", menu=vm)

        # BACKGROUND IMAGE
        bm = tk.Menu(menubar, tearoff=0)
        bm.add_command(label="Učitaj PNG pozadinu...",  command=self.bg_load)
        bm.add_command(label="Ukloni pozadinu",         command=self.bg_remove)
        bm.add_checkbutton(label="Prikaži / sakrij",
                           variable=self.bg_visible,
                           command=self.refresh_canvas)
        bm.add_separator()
        bm.add_command(label="Kalibracija (2 točke)...", command=self.bg_calibrate_start)
        bm.add_command(label="Poništi kalibraciju",      command=self.bg_calibrate_reset)
        bm.add_separator()
        bm.add_command(label="Transparentnost pozadine...", command=self.bg_alpha_dialog)
        menubar.add_cascade(label="Pozadina", menu=bm)

        # HELP
        hm = tk.Menu(menubar, tearoff=0)
        hm.add_command(label="About", command=lambda: messagebox.showinfo(
            "About",
            "Mini CAD\n\n"
            "Elementi: Linija, Kružnica, Pravokutnik,\n"
            "          Točka, Luk, Tekst, Kota\n\n"
            "Formati:  CSV, DXF, SVG, XLSX, PNG\n\n"
            "Ctrl+C:   Kopiraj nacrt kao PNG\n"
            "Ctrl+Z:   Undo"))
        menubar.add_cascade(label="Help", menu=hm)

        self.root.config(menu=menubar)

    def create_canvas(self):
        self.fig, self.ax = plt.subplots(figsize=(6, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.refresh_canvas()
        self.root.bind("<Control-z>", lambda e: self.undo())
        self.root.bind("<Control-Z>", lambda e: self.undo())
        self.root.bind("<Control-c>", lambda e: self.copy_png_clipboard())
        self.root.bind("<Control-C>", lambda e: self.copy_png_clipboard())
        self.canvas.mpl_connect("motion_notify_event",  self.on_mouse_move)
        self.canvas.mpl_connect("button_press_event",   self.on_mouse_press)
        self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.canvas.mpl_connect("scroll_event",         self.on_scroll)

    def create_statusbar(self):
        frm = tk.Frame(self.root, bd=1, relief=tk.SUNKEN, bg="#f0f0f0")
        frm.pack(side=tk.BOTTOM, fill=tk.X)
        self.status_var = tk.StringVar(value="X: —          Y: —")
        tk.Label(frm, textvariable=self.status_var,
                 font=("Courier", 10), bg="#f0f0f0",
                 anchor=tk.W, padx=8).pack(side=tk.LEFT)
        self.snap_status_var = tk.StringVar(value="")
        tk.Label(frm, textvariable=self.snap_status_var,
                 font=("Courier", 10), bg="#f0f0f0", fg="#0055aa",
                 anchor=tk.E, padx=8).pack(side=tk.RIGHT)

    def on_mouse_move(self, event):
        if event.xdata is not None and event.ydata is not None:
            x, y = event.xdata, event.ydata
            if self.snap_enabled.get():
                xs, ys = self._snap(x, y)
                self.status_var.set(
                    f"X: {x:9.3f}   Y: {y:9.3f}     "
                    f"snap → {xs:.2f}, {ys:.2f}")
            else:
                self.status_var.set(f"X: {x:10.3f}      Y: {y:10.3f}")
        else:
            self.status_var.set("X: —          Y: —")

        if self._pan_active and self._pan_inv is not None \
                and event.x is not None:
            x0, y0 = self._pan_inv.transform(
                (self._pan_press_xy[0], self._pan_press_xy[1]))
            x1, y1 = self._pan_inv.transform((event.x, event.y))
            dx, dy = x1 - x0, y1 - y0
            xl = (self._pan_xlim[0] - dx, self._pan_xlim[1] - dx)
            yl = (self._pan_ylim[0] - dy, self._pan_ylim[1] - dy)
            self._custom_xlim = xl
            self._custom_ylim = yl
            self.ax.set_xlim(*xl)
            self.ax.set_ylim(*yl)
            self.canvas.draw_idle()

    def on_mouse_press(self, event):
        if event.button == 3 and event.inaxes == self.ax:
            self._pan_active   = True
            self._pan_press_xy = (event.x, event.y)
            self._pan_xlim     = self.ax.get_xlim()
            self._pan_ylim     = self.ax.get_ylim()
            self._pan_inv      = self.ax.transData.inverted()

    def on_mouse_release(self, event):
        if event.button == 3:
            self._pan_active = False

    def on_scroll(self, event):
        if event.inaxes != self.ax or event.xdata is None:
            return
        factor = 0.85 if event.button == 'up' else 1.0 / 0.85
        xc, yc = event.xdata, event.ydata
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        self._custom_xlim = (xc + (xlim[0]-xc)*factor,
                             xc + (xlim[1]-xc)*factor)
        self._custom_ylim = (yc + (ylim[0]-yc)*factor,
                             yc + (ylim[1]-yc)*factor)
        self.ax.set_xlim(*self._custom_xlim)
        self.ax.set_ylim(*self._custom_ylim)
        self.canvas.draw_idle()

    # ----------------------------------------
    # Snap
    # ----------------------------------------

    def _snap(self, x, y):
        g = self.snap_grid
        return round(round(x/g)*g, 6), round(round(y/g)*g, 6)

    def _snap_toggle(self):
        if self.snap_enabled.get():
            self.snap_status_var.set(f"SNAP  grid={self.snap_grid}")
        else:
            self.snap_status_var.set("")

    def _set_snap_grid(self, v):
        self.snap_grid = v
        if self.snap_enabled.get():
            self.snap_status_var.set(f"SNAP  grid={self.snap_grid}")

    def _sv(self, val):
        """Primjeni snap na jednu koordinatu ako je snap uključen."""
        if self.snap_enabled.get():
            g = self.snap_grid
            return round(round(val/g)*g, 6)
        return val

    # ----------------------------------------
    # New / Undo / Delete
    # ----------------------------------------

    def new_drawing(self):
        all_empty = not any([self.lines, self.circles, self.rectangles,
                              self.points, self.arcs, self.texts, self.dimensions])
        if not all_empty:
            if not messagebox.askyesno("New", "Obrisati sve elemente i početi novi crtež?"):
                return
        self._clear_all()
        self.refresh_canvas()

    def undo(self):
        if not self.history:
            messagebox.showinfo("Undo", "Nema više elemenata za poništiti.")
            return
        kind, obj = self.history.pop()
        lst = self._list_for(kind)
        if lst is not None and obj in lst:
            lst.remove(obj)
        self.refresh_canvas()

    def _list_for(self, kind):
        return {"line": self.lines, "circle": self.circles,
                "rectangle": self.rectangles, "point": self.points,
                "arc": self.arcs, "text": self.texts,
                "dimension": self.dimensions}.get(kind)

    def delete_dialog(self):
        items = []
        for i, o in enumerate(self.lines):
            items.append(("line", i,
                f"Linija {i+1}:  ({o.x1},{o.y1}) → ({o.x2},{o.y2})"))
        for i, o in enumerate(self.circles):
            items.append(("circle", i,
                f"Kružnica {i+1}:  c=({o.cx},{o.cy})  r={o.r}"))
        for i, o in enumerate(self.rectangles):
            items.append(("rectangle", i,
                f"Pravokutnik {i+1}:  ({o.x1},{o.y1})–({o.x2},{o.y2})"))
        for i, o in enumerate(self.points):
            items.append(("point", i, f"Točka {i+1}:  ({o.x},{o.y})"))
        for i, o in enumerate(self.arcs):
            items.append(("arc", i,
                f"Luk {i+1}:  c=({o.cx},{o.cy})  r={o.r}  {o.angle1}°–{o.angle2}°"))
        for i, o in enumerate(self.texts):
            items.append(("text", i,
                f"Tekst {i+1}:  ({o.x},{o.y})  \"{o.content}\""))
        for i, o in enumerate(self.dimensions):
            items.append(("dimension", i,
                f"Kota {i+1}:  ({o.x1},{o.y1})–({o.x2},{o.y2})  off={o.offset}"))

        if not items:
            messagebox.showinfo("Brisanje", "Nema elemenata za brisanje.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Obriši element")
        dlg.grab_set(); dlg.resizable(False, False)
        tk.Label(dlg, text="Odaberi element koji želiš obrisati:",
                 font=("", 10, "bold")).pack(padx=12, pady=(10, 4))
        frm = tk.Frame(dlg)
        frm.pack(padx=12, pady=4, fill=tk.BOTH, expand=True)
        sb = tk.Scrollbar(frm, orient=tk.VERTICAL)
        lb = tk.Listbox(frm, yscrollcommand=sb.set, width=64,
                        height=min(len(items), 12),
                        selectmode=tk.SINGLE, font=("Courier", 9))
        sb.config(command=lb.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        for _, _, opis in items:
            lb.insert(tk.END, opis)

        def potvrdi():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("Upozorenje",
                    "Nije odabran nijedan element.", parent=dlg)
                return
            kind, idx, _ = items[sel[0]]
            lst = self._list_for(kind)
            if lst:
                obj = lst.pop(idx)
                if (kind, obj) in self.history:
                    self.history.remove((kind, obj))
            dlg.destroy()
            self.refresh_canvas()

        bf = tk.Frame(dlg); bf.pack(pady=10)
        tk.Button(bf, text="Obriši", width=12, bg="#e05050", fg="white",
                  command=potvrdi).pack(side=tk.LEFT, padx=6)
        tk.Button(bf, text="Odustani", width=12,
                  command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        lb.bind("<Double-Button-1>", lambda e: potvrdi())

    def set_range(self, r):
        self.axis_range   = r
        self._custom_xlim = None
        self._custom_ylim = None
        self.refresh_canvas()

    def auto_range(self):
        self.axis_range   = None
        self._custom_xlim = None
        self._custom_ylim = None
        self.refresh_canvas()

    def reset_view(self):
        """Vraća pogled na zadnji odabrani raspon, poništava pan/zoom."""
        self._custom_xlim = None
        self._custom_ylim = None
        self.refresh_canvas()

    # ----------------------------------------
    # Dijalozi za unos
    # ----------------------------------------

    def ask_values(self, fields, with_linetype=False):
        dlg = tk.Toplevel(self.root)
        dlg.title("Unos"); dlg.grab_set()
        self._dialog_result = None
        entries = {}
        for i, f in enumerate(fields):
            tk.Label(dlg, text=f, width=22, anchor=tk.W).grid(
                row=i, column=0, padx=8, pady=4)
            e = tk.Entry(dlg, width=14)
            e.grid(row=i, column=1, padx=8, pady=4)
            entries[f] = e
        list(entries.values())[0].focus_set()

        lt_var = None
        if with_linetype:
            n = len(fields)
            tk.Label(dlg, text="linetype", width=22, anchor=tk.W).grid(
                row=n, column=0, padx=8, pady=4)
            lt_var = tk.StringVar(value='CONTINUOUS')
            om = tk.OptionMenu(dlg, lt_var, *LINETYPES)
            om.config(width=14)
            om.grid(row=n, column=1, padx=8, pady=4, sticky=tk.W)

        def submit(event=None):
            try:
                vals = [float(entries[f].get()) for f in fields]
                coord_keys = {"x","y","x1","y1","x2","y2","cx","cy"}
                snapped = [self._sv(v) if fields[i].split()[0].lower() in coord_keys
                           else v for i, v in enumerate(vals)]
                if with_linetype:
                    snapped.append(lt_var.get())
                self._dialog_result = snapped
                dlg.destroy()
            except ValueError:
                messagebox.showerror("Greška", "Unesite ispravne brojeve!", parent=dlg)

        btn_row = len(fields) + (1 if with_linetype else 0)
        tk.Button(dlg, text="OK", command=submit).grid(
            row=btn_row, column=0, columnspan=2, pady=8)
        dlg.bind("<Return>", submit)
        dlg.wait_window()
        return self._dialog_result

    def ask_text_values(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Dodaj tekst"); dlg.grab_set()
        self._dialog_result = None
        labels = ["x", "y", "veličina fonta (pt)", "rotacija (°)"]
        entries = {}
        for i, f in enumerate(labels):
            tk.Label(dlg, text=f, width=20, anchor=tk.W).grid(
                row=i, column=0, padx=8, pady=4)
            e = tk.Entry(dlg, width=14)
            e.grid(row=i, column=1, padx=8, pady=4)
            entries[f] = e
        entries["veličina fonta (pt)"].insert(0, "10")
        entries["rotacija (°)"].insert(0, "0")
        n = len(labels)
        tk.Label(dlg, text="tekst", width=20, anchor=tk.W).grid(
            row=n, column=0, padx=8, pady=4)
        txt_e = tk.Entry(dlg, width=28)
        txt_e.grid(row=n, column=1, padx=8, pady=4)
        tk.Label(dlg, text="linetype", width=20, anchor=tk.W).grid(
            row=n+1, column=0, padx=8, pady=4)
        lt_var = tk.StringVar(value='CONTINUOUS')
        om = tk.OptionMenu(dlg, lt_var, *LINETYPES)
        om.config(width=14)
        om.grid(row=n+1, column=1, padx=8, pady=4, sticky=tk.W)
        list(entries.values())[0].focus_set()

        def submit(event=None):
            try:
                x   = self._sv(float(entries["x"].get()))
                y   = self._sv(float(entries["y"].get()))
                fs  = float(entries["veličina fonta (pt)"].get())
                rot = float(entries["rotacija (°)"].get())
                content = txt_e.get().strip()
                if not content:
                    messagebox.showerror("Greška", "Unesite tekst!", parent=dlg)
                    return
                self._dialog_result = (x, y, content, fs, rot, 0.0, lt_var.get())
                dlg.destroy()
            except ValueError:
                messagebox.showerror("Greška", "Unesite ispravne koordinate!", parent=dlg)

        tk.Button(dlg, text="OK", command=submit).grid(
            row=n+2, column=0, columnspan=2, pady=10)
        dlg.bind("<Return>", submit)
        dlg.wait_window()
        return self._dialog_result

    def _add(self, kind, obj):
        self._list_for(kind).append(obj)
        self.history.append((kind, obj))
        self.refresh_canvas()

    def add_line_dialog(self):
        v = self.ask_values(["x1", "y1", "x2", "y2"], with_linetype=True)
        if v: self._add("line", Line(*v))

    def add_circle_dialog(self):
        v = self.ask_values(["cx", "cy", "r"], with_linetype=True)
        if v: self._add("circle", Circle(*v))

    def add_rectangle_dialog(self):
        v = self.ask_values(["x1", "y1", "x2", "y2"], with_linetype=True)
        if v: self._add("rectangle", Rectangle(*v))

    def add_point_dialog(self):
        v = self.ask_values(["x", "y"], with_linetype=True)
        if v: self._add("point", Point(*v))

    def add_arc_dialog(self):
        v = self.ask_values(["cx", "cy", "r",
                              "kut početka (°)", "kut kraja (°)"], with_linetype=True)
        if v: self._add("arc", Arc(*v))

    def add_text_dialog(self):
        v = self.ask_text_values()
        if v: self._add("text", Text(*v))

    def add_dimension_dialog(self):
        v = self.ask_values(["x1", "y1", "x2", "y2",
                              "pomak kotne linije (offset)"], with_linetype=True)
        if v: self._add("dimension", Dimension(*v))

    # ----------------------------------------
    # Crtanje
    # ----------------------------------------

    def refresh_canvas(self):
        self.ax.clear()
        self.ax.axhline(0, color='black', linewidth=0.8)
        self.ax.axvline(0, color='black', linewidth=0.8)
        self.ax.grid(True, linestyle='--', alpha=0.3)
        self.ax.set_aspect('equal')
        if self._custom_xlim is not None:
            self.ax.set_xlim(*self._custom_xlim)
            self.ax.set_ylim(*self._custom_ylim)
        elif self.axis_range is not None:
            self.ax.set_xlim(-self.axis_range*0.05, self.axis_range)
            self.ax.set_ylim(-self.axis_range*0.05, self.axis_range)

        # Pozadinska slika
        if self.bg_image is not None and self.bg_visible.get():
            self._draw_bg_image()

        for o in self.lines:
            self.ax.plot([o.x1,o.x2],[o.y1,o.y2],
                         color='blue', linestyle=_ls(o.linetype))
        for o in self.circles:
            self.ax.add_patch(plt.Circle((o.cx,o.cy),o.r,
                                         fill=False, color='red',
                                         linestyle=_ls(o.linetype)))
        for o in self.rectangles:
            x=min(o.x1,o.x2); y=min(o.y1,o.y2)
            w=abs(o.x2-o.x1); h=abs(o.y2-o.y1)
            self.ax.add_patch(plt.Rectangle((x,y),w,h,
                                             fill=False, color='green',
                                             linestyle=_ls(o.linetype)))
        for o in self.points:
            self.ax.plot(o.x, o.y, 'ko', markersize=4)
            self.ax.annotate(f"({o.x},{o.y})", (o.x,o.y),
                             textcoords="offset points", xytext=(5,5),
                             fontsize=7, color='dimgray')
        for o in self.arcs:
            self.ax.add_patch(mpatches.Arc(
                (o.cx,o.cy), 2*o.r, 2*o.r,
                angle=0, theta1=o.angle1, theta2=o.angle2,
                color='darkorange', linewidth=1.5,
                linestyle=_ls(o.linetype)))
        for o in self.texts:
            self.ax.text(o.x, o.y, o.content,
                         fontsize=o.fontsize, color='purple', va='bottom',
                         rotation=o.rotation, rotation_mode='anchor')
        for o in self.dimensions:
            self._draw_dimension(o)

        self.canvas.draw()

    # ----------------------------------------
    # Background image
    # ----------------------------------------

    def _draw_bg_image(self):
        """Prikazuje pozadinsku sliku u koordinatnom sustavu nacrta."""
        import numpy as np
        img = self.bg_image          # PIL Image
        w_px, h_px = img.size       # dimenzije u pikselima

        if self.cal_scale is not None:
            # Kalibrirana slika: znamo skalu i offset
            s = self.cal_scale
            ox, oy = self.cal_offset
            # extent: [left, right, bottom, top] u CAD koordinatama
            extent = [ox, ox + w_px * s, oy, oy + h_px * s]
        else:
            # Nekalibrirana: prikaži u rasponu osi (0..axis_range)
            r = self.axis_range if self.axis_range else 100
            extent = [0, r, 0, r]

        self.ax.imshow(
            img,
            extent=extent,
            origin='upper',          # PNG (0,0) je gore lijevo
            aspect='auto',
            alpha=self.bg_alpha.get(),
            zorder=0)                # ispod svega

    def bg_load(self):
        """Učitava PNG datoteku kao pozadinu."""
        from PIL import Image
        fn = filedialog.askopenfilename(
            title="Odaberi pozadinsku sliku",
            filetypes=[("PNG slike", "*.png"),
                       ("Sve slike", "*.png *.jpg *.jpeg *.bmp *.tif")])
        if not fn: return
        try:
            self.bg_image = Image.open(fn).convert("RGBA")
            self.cal_scale  = None
            self.cal_offset = (0.0, 0.0)
            self.cal_points = []
            self.refresh_canvas()
            messagebox.showinfo("Pozadina",
                f"Slika učitana: {fn}\n\n"
                "Slika je prikazana nekalibrirana.\n"
                "Koristi Pozadina → Kalibracija za postavljanje mjerila.")
        except Exception as e:
            messagebox.showerror("Greška", f"Ne mogu učitati sliku:\n{e}")

    def bg_remove(self):
        """Uklanja pozadinsku sliku."""
        self.bg_image   = None
        self.cal_scale  = None
        self.cal_offset = (0.0, 0.0)
        self.cal_points = []
        self.refresh_canvas()

    def bg_alpha_dialog(self):
        """Dijalog za podešavanje transparentnosti pozadine."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Transparentnost pozadine")
        dlg.grab_set(); dlg.resizable(False, False)

        tk.Label(dlg, text="Transparentnost (0.0 = nevidljivo, 1.0 = puno):",
                 padx=12).pack(pady=(12, 4))

        scale = tk.Scale(dlg, from_=0.0, to=1.0, resolution=0.05,
                         orient=tk.HORIZONTAL, length=260,
                         variable=self.bg_alpha,
                         command=lambda v: self.refresh_canvas())
        scale.pack(padx=12, pady=4)

        tk.Button(dlg, text="Zatvori", command=dlg.destroy).pack(pady=10)

    def bg_calibrate_reset(self):
        """Poništava kalibraciju — slika se vraća na nekalibriran prikaz."""
        self.cal_scale  = None
        self.cal_offset = (0.0, 0.0)
        self.cal_points = []
        self._cal_end_mode()
        self.refresh_canvas()
        messagebox.showinfo("Kalibracija", "Kalibracija poništena.")

    def bg_calibrate_start(self):
        """Pokreće mod kalibracije: korisnik klikne 2 točke na slici."""
        if self.bg_image is None:
            messagebox.showwarning("Kalibracija",
                "Prvo učitaj pozadinsku sliku\n(Pozadina → Učitaj PNG pozadinu).")
            return

        self.cal_points = []
        self._cal_markers_clear()
        self.cal_mode = True

        # Povežemo klik na canvas
        if self._cal_click_cid:
            self.canvas.mpl_disconnect(self._cal_click_cid)
        self._cal_click_cid = self.canvas.mpl_connect(
            'button_press_event', self._cal_on_click)

        messagebox.showinfo("Kalibracija — korak 1/2",
            "Klikni PRVU točku na slici (npr. početak poznate kote).\n\n"
            "Nakon klika upisat ćeš drugu točku i stvarnu mjeru.")

    def _cal_on_click(self, event):
        """Prima klikove za kalibraciju (max 2)."""
        if event.inaxes != self.ax: return
        if event.button != 1: return        # samo lijevi klik

        x, y = event.xdata, event.ydata
        self.cal_points.append((x, y))

        # Nacrtaj marker
        m, = self.ax.plot(x, y, 'r+', markersize=14, markeredgewidth=2, zorder=10)
        self.cal_markers.append(m)
        self.canvas.draw()

        if len(self.cal_points) == 1:
            messagebox.showinfo("Kalibracija — korak 2/2",
                f"Prva točka: ({x:.3f}, {y:.3f})\n\n"
                "Klikni DRUGU točku na slici (npr. kraj iste kote).")

        elif len(self.cal_points) == 2:
            self._cal_end_mode()
            self._cal_ask_real_distance()

    def _cal_end_mode(self):
        if self._cal_click_cid:
            self.canvas.mpl_disconnect(self._cal_click_cid)
            self._cal_click_cid = None
        self.cal_mode = False

    def _cal_markers_clear(self):
        for m in self.cal_markers:
            try: m.remove()
            except Exception: pass
        self.cal_markers = []

    def _cal_ask_real_distance(self):
        """Dijalog za unos stvarne mjere i postavljanje ishodišta."""
        p1, p2 = self.cal_points
        # udaljenost u ax koordinatama (piksel koordinate nekalibriranog prikaza)
        r = self.axis_range if self.axis_range else 100
        w_px, h_px = self.bg_image.size

        # ax koordinata → piksel (nekalibrirana slika pokriva 0..r u oba smjera)
        ax_to_px_x = w_px / r
        ax_to_px_y = h_px / r

        # udaljenost u pikselima
        dpx = (p2[0]-p1[0]) * ax_to_px_x
        dpy = (p2[1]-p1[1]) * ax_to_px_y
        dist_px = math.hypot(dpx, dpy)

        dlg = tk.Toplevel(self.root)
        dlg.title("Kalibracija — stvarna mjera")
        dlg.grab_set(); dlg.resizable(False, False)

        tk.Label(dlg,
            text=f"Kliknute točke:\n"
                 f"  T1: ({p1[0]:.2f}, {p1[1]:.2f})\n"
                 f"  T2: ({p2[0]:.2f}, {p2[1]:.2f})\n\n"
                 f"Udaljenost na slici: {dist_px:.1f} px\n\n"
                 "Unesi stvarnu mjeru između tih točaka:",
            justify=tk.LEFT, padx=14, pady=8).grid(
            row=0, column=0, columnspan=2)

        tk.Label(dlg, text="Stvarna mjera (mm / m / bilo što):",
                 padx=12).grid(row=1, column=0, sticky=tk.W)
        e_dist = tk.Entry(dlg, width=14)
        e_dist.grid(row=1, column=1, padx=8, pady=4)
        e_dist.focus_set()

        tk.Label(dlg,
            text="CAD koordinate prve točke\n(što treba biti T1 u nacrtu):",
            padx=12, justify=tk.LEFT).grid(row=2, column=0, sticky=tk.W, pady=(8,0))

        tk.Label(dlg, text="X ishodišta (T1):", padx=12).grid(
            row=3, column=0, sticky=tk.W)
        e_x = tk.Entry(dlg, width=14)
        e_x.insert(0, "0")
        e_x.grid(row=3, column=1, padx=8, pady=2)

        tk.Label(dlg, text="Y ishodišta (T1):", padx=12).grid(
            row=4, column=0, sticky=tk.W)
        e_y = tk.Entry(dlg, width=14)
        e_y.insert(0, "0")
        e_y.grid(row=4, column=1, padx=8, pady=2)

        def apply_cal(event=None):
            try:
                real_dist = float(e_dist.get())
                cad_x1    = float(e_x.get())
                cad_y1    = float(e_y.get())
            except ValueError:
                messagebox.showerror("Greška", "Unesite ispravne brojeve!", parent=dlg)
                return
            if dist_px == 0:
                messagebox.showerror("Greška", "Točke su identične!", parent=dlg)
                return

            # skala: koliko CAD jedinica po pikselu
            scale = real_dist / dist_px

            # T1 u piksel koordinatama (od gornjeg lijevog ruba slike)
            px1 = p1[0] * ax_to_px_x
            py1 = (r - p1[1]) * ax_to_px_y   # flip Y (ax Y raste gore)

            # Offset: CAD (0,0) = piksel (0,0) slike, ali T1 mora biti cad_x1,cad_y1
            # ox = cad_x1 - px1 * scale
            # oy = cad_y1 - (h_px - py1) * scale   (bottom-left origin)
            ox = cad_x1 - px1 * scale
            oy = cad_y1 - (h_px - py1) * scale

            self.cal_scale  = scale
            self.cal_offset = (ox, oy)
            dlg.destroy()
            self._cal_markers_clear()
            self.refresh_canvas()
            messagebox.showinfo("Kalibracija",
                f"✓ Kalibracija uspješna!\n\n"
                f"Mjerilo: 1 px = {scale:.4f} jed.\n"
                f"Offset:  X={ox:.3f}  Y={oy:.3f}\n\n"
                "Sada crtaj po slici — koordinate odgovaraju nacrtu.")

        bf = tk.Frame(dlg); bf.grid(row=5, column=0, columnspan=2, pady=10)
        tk.Button(bf, text="Primijeni", width=12,
                  bg="#4070c0", fg="white",
                  command=apply_cal).pack(side=tk.LEFT, padx=6)
        tk.Button(bf, text="Odustani", width=12,
                  command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        dlg.bind("<Return>", apply_cal)

    def _draw_dimension(self, d):
        ax = self.ax
        x1,y1,x2,y2 = d.x1,d.y1,d.x2,d.y2
        off = d.offset
        dx=x2-x1; dy=y2-y1
        length = math.hypot(dx,dy)
        if length == 0: return
        nx=-dy/length; ny=dx/length
        px1,py1 = x1+nx*off, y1+ny*off
        px2,py2 = x2+nx*off, y2+ny*off
        kw = dict(color='#555555', linewidth=0.8)
        # kotna linija s dvostrano strelicama
        ax.annotate("", xy=(px2,py2), xytext=(px1,py1),
                    arrowprops=dict(arrowstyle="<->", color='#555555',
                                    lw=0.8, mutation_scale=10))
        # svjedočnice
        ext = abs(off)*0.15
        ax.plot([x1, px1+nx*ext],[y1, py1+ny*ext], **kw)
        ax.plot([x2, px2+nx*ext],[y2, py2+ny*ext], **kw)
        # mjera
        mx,my = (px1+px2)/2, (py1+py2)/2
        angle_deg = math.degrees(math.atan2(dy,dx))
        if angle_deg > 90 or angle_deg < -90:
            angle_deg += 180
        ax.text(mx, my, f"{length:.2f}",
                ha='center', va='bottom', fontsize=8, color='#333333',
                rotation=angle_deg, rotation_mode='anchor')

    # ----------------------------------------
    # SVG  (zajednički generator)
    # ----------------------------------------

    def _generate_svg_text(self):
        """Generira ispravan SVG string.
        KLJUČNI FIX za mirror: SVG Y os raste prema dolje, matplotlib prema gore.
        Koristimo transform="scale(1,-1)" na grupi i kompenziramo translate."""
        r = self.axis_range if self.axis_range else 100
        margin = r * 0.05
        W = r + margin        # širina viewBoxa
        H = r + margin        # visina viewBoxa

        sw = max(W * 0.003, 0.3)   # stroke-width proporcionalan nacrtu

        out = []
        out.append(f'<svg xmlns="http://www.w3.org/2000/svg" '
                   f'width="{W:.1f}" height="{H:.1f}" '
                   f'viewBox="{-margin:.2f} {-margin:.2f} {W:.2f} {H:.2f}">')

        # Grupa s flipom: zrcalimo Y i translateamo natrag na pravo mjesto
        out.append(f'<g transform="translate(0,{r:.4f}) scale(1,-1)">')

        for o in self.lines:
            out.append(f'<line x1="{o.x1}" y1="{o.y1}" '
                       f'x2="{o.x2}" y2="{o.y2}" '
                       f'stroke="blue" stroke-width="{sw}"/>')
        for o in self.circles:
            out.append(f'<circle cx="{o.cx}" cy="{o.cy}" r="{o.r}" '
                       f'stroke="red" fill="none" stroke-width="{sw}"/>')
        for o in self.rectangles:
            x=min(o.x1,o.x2); y=min(o.y1,o.y2)
            w=abs(o.x2-o.x1); h=abs(o.y2-o.y1)
            out.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" '
                       f'stroke="green" fill="none" stroke-width="{sw}"/>')
        for o in self.points:
            out.append(f'<circle cx="{o.x}" cy="{o.y}" r="{sw*1.5:.3f}" '
                       f'fill="black"/>')
        for o in self.arcs:
            out.append(f'<path d="{self._arc_path(o)}" '
                       f'stroke="darkorange" fill="none" stroke-width="{sw}"/>')
        for o in self.dimensions:
            out.extend(self._dim_svg(o, sw))

        # Tekst: mora se re-flipati jer smo već flipali grupu
        for o in self.texts:
            # scale(1,-1) oko točke (o.x, o.y) vraća tekst u normalan položaj
            out.append(f'<text x="{o.x}" y="{o.y}" '
                       f'transform="scale(1,-1) translate(0,{-2*o.y:.4f})" '
                       f'font-size="{o.fontsize}" fill="purple">'
                       f'{o.content}</text>')

        out.append('</g>')
        out.append('</svg>')
        return "\n".join(out)

    def _arc_path(self, a):
        r1 = math.radians(a.angle1); r2 = math.radians(a.angle2)
        sx = a.cx + a.r*math.cos(r1); sy = a.cy + a.r*math.sin(r1)
        ex = a.cx + a.r*math.cos(r2); ey = a.cy + a.r*math.sin(r2)
        large = 1 if (a.angle2 - a.angle1) % 360 > 180 else 0
        return f"M {sx:.4f} {sy:.4f} A {a.r} {a.r} 0 {large} 1 {ex:.4f} {ey:.4f}"

    def _dim_svg(self, d, sw):
        """SVG elementi za kotnu liniju."""
        elems = []
        x1,y1,x2,y2 = d.x1,d.y1,d.x2,d.y2
        off = d.offset
        dx=x2-x1; dy=y2-y1
        length = math.hypot(dx,dy)
        if length == 0: return elems
        nx=-dy/length; ny=dx/length
        px1,py1 = x1+nx*off, y1+ny*off
        px2,py2 = x2+nx*off, y2+ny*off
        c = "#555555"
        # svjedočnice
        ext = abs(off)*0.15
        for ox,oy,fx,fy in [(x1,y1,px1+nx*ext,py1+ny*ext),
                             (x2,y2,px2+nx*ext,py2+ny*ext)]:
            elems.append(f'<line x1="{ox:.3f}" y1="{oy:.3f}" '
                         f'x2="{fx:.3f}" y2="{fy:.3f}" '
                         f'stroke="{c}" stroke-width="{sw}"/>')
        # kotna linija
        elems.append(f'<line x1="{px1:.3f}" y1="{py1:.3f}" '
                     f'x2="{px2:.3f}" y2="{py2:.3f}" '
                     f'stroke="{c}" stroke-width="{sw}" '
                     f'marker-start="url(#a)" marker-end="url(#a)"/>')
        # mjera (re-flip teksta)
        mx,my = (px1+px2)/2, (py1+py2)/2
        angle_deg = math.degrees(math.atan2(dy,dx))
        if angle_deg > 90 or angle_deg < -90: angle_deg += 180
        fs = sw * 4
        elems.append(
            f'<text x="{mx:.3f}" y="{my:.3f}" '
            f'transform="scale(1,-1) translate(0,{-2*my:.3f}) '
            f'rotate({angle_deg:.2f},{mx:.3f},{-my:.3f})" '
            f'font-size="{fs:.3f}" fill="{c}" text-anchor="middle">'
            f'{length:.2f}</text>')
        return elems

    def save_svg(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".svg",
            filetypes=[("SVG datoteke", "*.svg")])
        if not filename: return
        with open(filename, "w", encoding="utf-8") as f:
            f.write(self._generate_svg_text())
        messagebox.showinfo("Spremljeno", f"SVG snimljen:\n{filename}")

    def copy_svg_clipboard(self):
        svg = self._generate_svg_text()
        self.root.clipboard_clear()
        self.root.clipboard_append(svg)
        self.root.update()
        messagebox.showinfo("Clipboard",
            "SVG kopiran u međuspremnik.\n"
            "Zalijepi ga u Inkscape, LibreOffice Draw,\n"
            "web editor ili bilo koji tekstualni editor.")

    # ----------------------------------------
    # CSV
    # ----------------------------------------

    def save_png(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG slike", "*.png")])
        if not filename: return
        self.fig.savefig(filename, dpi=150, bbox_inches='tight',
                         facecolor='white')
        messagebox.showinfo("Spremljeno", f"PNG snimljen:\n{filename}")

    def copy_png_clipboard(self):
        """Renderira nacrt u PNG i kopira u clipboard putem xclip.
        Zahtijeva: sudo apt install xclip"""
        import subprocess, tempfile, os, shutil

        if not shutil.which("xclip"):
            messagebox.showerror(
                "xclip nije instaliran",
                "Za kopiranje PNG-a u clipboard potreban je xclip.\n\n"
                "Instaliraj ga s:\n"
                "  sudo apt install xclip")
            return

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmpname = tmp.name

        try:
            self.fig.savefig(tmpname, dpi=150, bbox_inches='tight',
                             facecolor='white')
            with open(tmpname, "rb") as f:
                png_data = f.read()

            proc = subprocess.run(
                ["xclip", "-selection", "clipboard",
                 "-t", "image/png", "-i"],
                input=png_data,
                capture_output=True)

            if proc.returncode == 0:
                messagebox.showinfo(
                    "Clipboard",
                    "Nacrt kopiran kao PNG u međuspremnik.\n"
                    "Zalijepi ga u LibreOffice, GIMP, Word...")
            else:
                messagebox.showerror("Greška",
                    f"xclip greška:\n{proc.stderr.decode()}")
        finally:
            os.unlink(tmpname)

    def save_csv(self):
        fn = filedialog.asksaveasfilename(defaultextension=".csv",
             filetypes=[("CSV datoteke","*.csv")])
        if not fn: return
        with open(fn,"w",newline="",encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["type","x1","y1","x2","y2","r",
                         "angle1","angle2","content","fontsize","offset","linetype"])
            for o in self.lines:
                w.writerow(["line",o.x1,o.y1,o.x2,o.y2,"","","","","","",o.linetype])
            for o in self.circles:
                w.writerow(["circle",o.cx,o.cy,"","",o.r,"","","","","",o.linetype])
            for o in self.rectangles:
                w.writerow(["rectangle",o.x1,o.y1,o.x2,o.y2,"","","","","","",o.linetype])
            for o in self.points:
                w.writerow(["point",o.x,o.y,"","","","","","","","",o.linetype])
            for o in self.arcs:
                w.writerow(["arc",o.cx,o.cy,"","",o.r,
                             o.angle1,o.angle2,"","","",o.linetype])
            for o in self.texts:
                w.writerow(["text",o.x,o.y,"","","",
                             o.rotation,o.oblique,o.content,o.fontsize,"",o.linetype])
            for o in self.dimensions:
                w.writerow(["dimension",o.x1,o.y1,o.x2,o.y2,"","","","","",
                             o.offset,o.linetype])
        messagebox.showinfo("Spremljeno", f"CSV snimljen:\n{fn}")

    def load_csv(self):
        fn = filedialog.askopenfilename(filetypes=[("CSV datoteke","*.csv")])
        if not fn: return
        self._clear_all()
        with open(fn,"r",encoding="utf-8") as f:
            for row in csv.DictReader(f):
                self._load_row(row)
        self.refresh_canvas()

    def _fv(self, row, key):
        v = row.get(key,"")
        return float(v) if v not in ("", None) else 0.0

    def _lt(self, row):
        v = str(row.get("linetype","") or "").upper().strip()
        return v if v in LINETYPES else 'CONTINUOUS'

    def _load_row(self, row):
        t = row.get("type",""); f = self._fv
        if t == "line":
            self.lines.append(Line(f(row,"x1"),f(row,"y1"),
                                   f(row,"x2"),f(row,"y2"), self._lt(row)))
        elif t == "circle":
            self.circles.append(Circle(f(row,"x1"),f(row,"y1"),
                                        f(row,"r"), self._lt(row)))
        elif t == "rectangle":
            self.rectangles.append(Rectangle(f(row,"x1"),f(row,"y1"),
                                              f(row,"x2"),f(row,"y2"), self._lt(row)))
        elif t == "point":
            self.points.append(Point(f(row,"x1"),f(row,"y1"), self._lt(row)))
        elif t == "arc":
            self.arcs.append(Arc(f(row,"x1"),f(row,"y1"),f(row,"r"),
                                  f(row,"angle1"),f(row,"angle2"), self._lt(row)))
        elif t == "text":
            fs  = float(row["fontsize"]) \
                  if row.get("fontsize","") not in ("",None) else 10
            rot = float(row["angle1"]) \
                  if row.get("angle1","") not in ("",None) else 0.0
            obl = float(row["angle2"]) \
                  if row.get("angle2","") not in ("",None) else 0.0
            self.texts.append(Text(f(row,"x1"),f(row,"y1"),
                                   row.get("content",""), fs, rot, obl, self._lt(row)))
        elif t == "dimension":
            off = float(row["offset"]) \
                  if row.get("offset","") not in ("",None) else 5.0
            self.dimensions.append(Dimension(f(row,"x1"),f(row,"y1"),
                                              f(row,"x2"),f(row,"y2"),
                                              off, self._lt(row)))

    def _clear_all(self):
        for lst in [self.lines, self.circles, self.rectangles,
                    self.points, self.arcs, self.texts,
                    self.dimensions, self.history]:
            lst.clear()

    # ----------------------------------------
    # XLSX
    # ----------------------------------------

    def save_xlsx(self):
        if not XLSX_OK:
            messagebox.showerror("Greška","pip install openpyxl"); return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx",
             filetypes=[("Excel datoteke","*.xlsx")])
        if not fn: return
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MiniCad"
        hdr = ["type","x1","y1","x2","y2","r",
               "angle1","angle2","content","fontsize","offset","linetype"]
        ws.append(hdr)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center")
        for o in self.lines:
            ws.append(["line",o.x1,o.y1,o.x2,o.y2,"","","","","","",o.linetype])
        for o in self.circles:
            ws.append(["circle",o.cx,o.cy,"","",o.r,"","","","","",o.linetype])
        for o in self.rectangles:
            ws.append(["rectangle",o.x1,o.y1,o.x2,o.y2,"","","","","","",o.linetype])
        for o in self.points:
            ws.append(["point",o.x,o.y,"","","","","","","","",o.linetype])
        for o in self.arcs:
            ws.append(["arc",o.cx,o.cy,"","",o.r,o.angle1,o.angle2,"","","",o.linetype])
        for o in self.texts:
            ws.append(["text",o.x,o.y,"","","",
                        o.rotation,o.oblique,o.content,o.fontsize,"",o.linetype])
        for o in self.dimensions:
            ws.append(["dimension",o.x1,o.y1,o.x2,o.y2,
                        "","","","","",o.offset,o.linetype])
        for col in ws.columns:
            mw = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(mw+2,8)
        wb.save(fn)
        messagebox.showinfo("Spremljeno", f"XLSX snimljen:\n{fn}")

    def load_xlsx(self):
        if not XLSX_OK:
            messagebox.showerror("Greška","pip install openpyxl"); return
        fn = filedialog.askopenfilename(filetypes=[("Excel datoteke","*.xlsx")])
        if not fn: return
        self._clear_all()
        wb = openpyxl.load_workbook(fn); ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = {k:("" if v is None else v) for k,v in zip(headers,row)}
            self._load_row(rd)
        self.refresh_canvas()

    # ----------------------------------------
    # DXF
    # ----------------------------------------

    def save_dxf(self):
        fn = filedialog.asksaveasfilename(defaultextension=".dxf",
             filetypes=[("DXF datoteke","*.dxf")])
        if not fn: return
        doc = ezdxf.new(); msp = doc.modelspace()
        for o in self.lines:
            lt = _ensure_lt(doc, o.linetype)
            msp.add_line((o.x1,o.y1),(o.x2,o.y2),
                         dxfattribs={"linetype": lt} if lt != 'CONTINUOUS' else {})
        for o in self.circles:
            lt = _ensure_lt(doc, o.linetype)
            msp.add_circle((o.cx,o.cy),o.r,
                           dxfattribs={"linetype": lt} if lt != 'CONTINUOUS' else {})
        for o in self.rectangles:
            lt = _ensure_lt(doc, o.linetype)
            msp.add_lwpolyline(
                [(o.x1,o.y1),(o.x2,o.y1),(o.x2,o.y2),(o.x1,o.y2),(o.x1,o.y1)],
                close=True,
                dxfattribs={"linetype": lt} if lt != 'CONTINUOUS' else {})
        for o in self.points:
            msp.add_point((o.x,o.y))
        for o in self.arcs:
            lt = _ensure_lt(doc, o.linetype)
            msp.add_arc((o.cx,o.cy),o.r,o.angle1,o.angle2,
                        dxfattribs={"linetype": lt} if lt != 'CONTINUOUS' else {})
        for o in self.texts:
            attribs = {"insert": (o.x, o.y), "height": o.fontsize * 0.35}
            if o.rotation: attribs["rotation"] = o.rotation
            if o.oblique:  attribs["oblique"]  = o.oblique
            msp.add_text(o.content, dxfattribs=attribs)
        for o in self.dimensions:
            try:
                msp.add_linear_dim(
                    base=(o.x1, o.y1+o.offset),
                    p1=(o.x1,o.y1), p2=(o.x2,o.y2))
            except Exception:
                msp.add_line((o.x1,o.y1),(o.x2,o.y2))
        doc.saveas(fn)
        messagebox.showinfo("Spremljeno", f"DXF snimljen:\n{fn}")

    def load_dxf(self):
        fn = filedialog.askopenfilename(filetypes=[("DXF datoteke","*.dxf")])
        if not fn: return
        self._clear_all()
        doc = ezdxf.readfile(fn); msp = doc.modelspace()
        for e in msp:
            lt = (e.dxf.get('linetype','') or 'CONTINUOUS').upper()
            if lt not in LINETYPES: lt = 'CONTINUOUS'

            if e.dxftype() == "LINE":
                self.lines.append(Line(e.dxf.start.x, e.dxf.start.y,
                                       e.dxf.end.x,   e.dxf.end.y, lt))
            elif e.dxftype() == "CIRCLE":
                self.circles.append(Circle(e.dxf.center.x, e.dxf.center.y,
                                           e.dxf.radius, lt))
            elif e.dxftype() == "ARC":
                self.arcs.append(Arc(e.dxf.center.x, e.dxf.center.y,
                                     e.dxf.radius,
                                     e.dxf.start_angle, e.dxf.end_angle, lt))
            elif e.dxftype() == "POINT":
                self.points.append(Point(e.dxf.location.x,
                                         e.dxf.location.y, lt))
            elif e.dxftype() == "TEXT":
                try:
                    ins = e.dxf.insert
                    h   = getattr(e.dxf, 'height',   3.5)
                    rot = getattr(e.dxf, 'rotation',  0.0)
                    obl = getattr(e.dxf, 'oblique',   0.0)
                    self.texts.append(
                        Text(ins.x, ins.y, e.dxf.text,
                             round(h / 0.35), rot, obl, lt))
                except Exception:
                    pass
            elif e.dxftype() == "MTEXT":
                try:
                    ins  = e.dxf.insert
                    h    = getattr(e.dxf, 'char_height', 3.5)
                    rot  = getattr(e.dxf, 'rotation',    0.0)
                    plain = _mtext_plain(e.text)
                    fs    = round(h / 0.35)
                    sp    = h * 1.67
                    for i, line in enumerate(plain.split('\n')):
                        if line.strip():
                            self.texts.append(
                                Text(ins.x, ins.y - i * sp,
                                     line, fs, rot, 0.0, lt))
                except Exception:
                    pass
            elif e.dxftype() == "LWPOLYLINE":
                pts = list(e.get_points())
                n   = len(pts)
                for i in range(n - 1):
                    self.lines.append(Line(pts[i][0], pts[i][1],
                                           pts[i+1][0], pts[i+1][1], lt))
                if e.closed and n >= 2:
                    self.lines.append(Line(pts[-1][0], pts[-1][1],
                                           pts[0][0],  pts[0][1], lt))
            elif e.dxftype() == "SPLINE":
                pts = list(e.flattening(0.01))
                for i in range(len(pts) - 1):
                    self.lines.append(Line(pts[i][0], pts[i][1],
                                           pts[i+1][0], pts[i+1][1], lt))
            elif e.dxftype() == "POLYLINE":
                try:
                    vertices = list(e.vertices)
                    pts = [(v.dxf.location.x, v.dxf.location.y)
                           for v in vertices]
                    n = len(pts)
                    for i in range(n - 1):
                        self.lines.append(Line(pts[i][0], pts[i][1],
                                               pts[i+1][0], pts[i+1][1], lt))
                    if e.is_closed and n >= 2:
                        self.lines.append(Line(pts[-1][0], pts[-1][1],
                                               pts[0][0],  pts[0][1], lt))
                except Exception:
                    pass
        self.refresh_canvas()


# ----------------------------------------
# Pokretanje
# ----------------------------------------
root = tk.Tk()
app = CADApp(root)
root.mainloop()
